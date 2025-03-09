#############################################################
#           Activities Tracker by Konnor Mascara
# 
# Note: Parts have the code has been influenced by similar 
# projects this may also include similar styles and blocks 
# of code copied from the imported files. This is a side 
# project and should not be taken professionally.
#
# Description: Allows event tracking for each person on 3
# units with a monthly submission.
#
# Done   Name     Description
# ----   ----     ----------- 
# Yes  | Screen | Have proper screen setup and refreshing
# Yes  | Events | Have proper events checking
# Yes  | Add    | Allow for people to be added
# Yes  | Delete | Allow for people to be deleted
# Yes  | Save   | Allow for the day to be submitted
# Yes  | Month  | Allow for the month to be submitted
# Yes  | Review | Check for duplicates and do the totals * 
# No   | Old    | Be able to open old stuff
#
# * Duplicates is completed, totals was not.
#
#############################################################

import customtkinter #GUI
import pandas as pd #Data management
from openpyxl import load_workbook #For getting totals

#Global Vars and setup
customtkinter.set_appearance_mode("System")
customtkinter.set_default_color_theme("blue") 
df1 = pd.read_csv('app_files/U1.csv') #Move csv's to a dataframe
df2 = pd.read_csv('app_files/U2.csv')
df3 = pd.read_csv('app_files/U3.csv')
event = "" #keep track of event picked
act = {} #Keep track of the people

#unit picker is made
class Unitpicker(customtkinter.CTkToplevel):
    def __init__(self):
        super().__init__()
        global df1, df2, df3
        self.row_count = 0

        #Setup window
        self.title("Activities Tracker by KAM")
        self.after(250, lambda: self.iconbitmap('app_files/logo.ico'))

        #Determine if things in grid grows(1) or doesnt (0)
        self.grid_columnconfigure((0,1), weight=1)
        self.grid_rowconfigure((0,1,2), weight=0)
        self.grid_rowconfigure(3, weight=1)

        #Select names to add to event
        self.tabview = customtkinter.CTkTabview(self, width=250, )
        self.tabview.grid(row=3, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
        self.tabview.add("Unit 1")
        self.tabview.add("Unit 2")
        self.tabview.add("Unit 3")
        self.tabview.tab('Unit 1').columnconfigure((0,1,2), weight=1)
        self.tabview.tab('Unit 1').rowconfigure(0, weight=1)
        self.tabview.tab('Unit 2').rowconfigure(0, weight=1)
        self.tabview.tab('Unit 3').rowconfigure(0, weight=1)
        self.tabview.tab('Unit 2').columnconfigure((0,1,2), weight=1)
        self.tabview.tab('Unit 3').columnconfigure((0,1,2), weight=1)

        #Setup scrolling
        self.scroll1 = customtkinter.CTkScrollableFrame(self.tabview.tab('Unit 1'))
        self.scroll1.grid(row=0, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")
        self.scroll1.columnconfigure((0,1,2), weight=1)
        self.scroll1.rowconfigure(0, weight=1)
        self.scroll2 = customtkinter.CTkScrollableFrame(self.tabview.tab('Unit 2'))
        self.scroll2.grid(row=0, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")
        self.scroll2.columnconfigure((0,1,2), weight=1)
        self.scroll2.rowconfigure(0, weight=1)
        self.scroll3 = customtkinter.CTkScrollableFrame(self.tabview.tab('Unit 3'))
        self.scroll3.grid(row=0, column=0, columnspan=3, padx=10, pady=10, sticky="nsew")
        self.scroll3.columnconfigure((0,1,2), weight=1)
        self.scroll3.rowconfigure(0, weight=1)
        
        for col in df1.iloc[:,0]: #Put everyname in the unit on the page
                self.setup(col, self.scroll1)
                self.row_count += 2
        for col in df2.iloc[:,0]: #Put everyname in the unit on the page
                self.setup(col, self.scroll2)
                self.row_count += 2
        for col in df3.iloc[:,0]: #Put everyname in the unit on the page
                self.setup(col, self.scroll3)
                self.row_count += 2
    
    #Display names
    def setup(self, name, unit):
        global act, event

        self.delete = customtkinter.CTkRadioButton(unit, text="Delete")
        self.delete.configure(command=lambda : self.getrid(name))
        self.delete.grid(row=self.row_count, column=0, padx=10, pady=10)

        self.name_label = customtkinter.CTkLabel(unit, text=name, font=customtkinter.CTkFont(size=20, weight="bold"))
        self.name_label.grid(row=self.row_count, column=1, padx=10, pady=10)
        
        checkbox = customtkinter.CTkCheckBox(unit, text="Attended",command=lambda : self.complete_day(name))
        checkbox.grid(row=self.row_count, column=2, pady=10)

        self.line = customtkinter.CTkFrame(unit, height=2, fg_color="grey")
        self.line.grid(row=self.row_count+1, column=0, columnspan=3, sticky="ew")

        if name not in act: #Add to matrix
            act[name] = [0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        else: #Change checkbox to be checked
            things = ["visits","music","outdoor","religion","food/drink","cards/games","arts/crafts","physical games","group activity",
                  "bus trips","special event","act on the go","noted","misc"]
            num = things.index(event)
            change = act[name][num]
            if change == 1:
                checkbox.select()
            
    #update local directory on buttons hit
    def complete_day(self, name):
        global act, event
        things = ["visits","music","outdoor","religion","food/drink","cards/games","arts/crafts","physical games","group activity",
                  "bus trips","special event","act on the go","noted","misc"]
        num = things.index(event)
        change = act[name][num]
        if change == 1:
            act[name][num] = 0
        else:
            act[name][num] = 1

    #Delete person
    def getrid(self, name):
        global df1, df2, df3, act
        #df1 = df1.drop(df1[df1['name'] == name].index)
        df1 = df1[df1.iloc[:,0] != name]
        df2 = df2[df2.iloc[:,0] != name]
        df3 = df3[df3.iloc[:,0] != name]
        del act[name]
        print("--------afterd--------")
        print(act)
        print("\n")
        print(df1)
        print("--------afterd--------")

    #When exited delete window
    def deletion(self):
        self.destroy()

#The main window
class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        #Setup window
        self.title("Activities Tracker by KAM")
        self.resizable(False, False)
        self.iconbitmap("app_files/logo.ico")

        #Adding Options
        self.name_label = customtkinter.CTkLabel(self, text="Add a name then pick the unit", 
                                                   font=customtkinter.CTkFont(size=20, weight="bold"))
        self.name_label.grid(row=0, column=0, padx=10, pady=10, sticky='n')

        self.name_input = customtkinter.CTkEntry(self, placeholder_text="Name")
        self.name_input.grid(row=1, column=0, padx=10, pady=10, sticky='n')

        self.unit_menu = customtkinter.CTkOptionMenu(self, values=["U1","U2","U3"], command=self.addingperson)
        self.unit_menu.grid(row=2, column=0, padx=10, pady=10, sticky='n')
        self.unit_menu.set("Unit")

        self.line1 = customtkinter.CTkFrame(self, height=2, fg_color="white")
        self.line1.grid(row=3, column=0, padx = 10, pady = 10, sticky="ew")

        #Event Options
        self.event_label = customtkinter.CTkLabel(self,text="Pick a event to start", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.event_label.grid(row=4, column=0, padx=10, pady=10)

        self.eventselect_label = customtkinter.CTkOptionMenu(self, command=self.eventselect,
            values=["visits","music","outdoor","religion","food/drink","cards/games","arts/crafts","physical games","group activity",
                  "bus trips","special event","act on the go","noted","misc"])
        self.eventselect_label.grid(row=5, column=0, padx=10, pady=10)
        self.eventselect_label.set("Event")

        self.line2 = customtkinter.CTkFrame(self, height=2, fg_color="white")
        self.line2.grid(row=6, column=0, padx = 10, pady = 10, sticky="ew")

        #Saving Options
        self.day_label = customtkinter.CTkLabel(self, text="Pick a day to save selection", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.day_label.grid(row=7, column=0, padx=10, pady=10)

        self.day_menu = customtkinter.CTkOptionMenu(self, values=[str(i) for i in range(1,32)], command=self.submitday)
        self.day_menu.grid(row=8, column=0, padx=10, pady=10)
        self.day_menu.set("Day")

        self.line3 = customtkinter.CTkFrame(self, height=2, fg_color="white")
        self.line3.grid(row=9, column=0, padx = 10, pady = 10, sticky="ew")

        #Excel Options
        self.month_label = customtkinter.CTkLabel(self,text="Pick a month to finish",
                                                  font=customtkinter.CTkFont(size=20, weight="bold"))
        self.month_label.grid(row=10, column=0, padx=10, pady=10)

        self.month_menu = customtkinter.CTkOptionMenu(self, command=self.submitmonth,
            values=["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"])
        self.month_menu.grid(row=11, column=0, padx=10, pady=10)
        self.month_menu.set("Month")

        #Add a person
    
    #Adds person
    def addingperson(self, unit: str):
        global df1, df2, df3, act
        self.unit_menu.set("Unit")
        name = self.name_input.get()
        self.name_input.delete(0,"end")
        newrow = [[name,None,None,None,None,None,None,None,None,None,None,None,None,None,None]]

        if name: #If no name do nothing
            if unit == "U1" and (name not in df1.iloc[:,0].values):
                colNames = df1.columns #get columns
                df_new = pd.DataFrame(data=newrow, columns=colNames) #Create new with columns
                df_together = pd.concat([df1,df_new],axis=0) #put together
                df1 = df_together.sort_values(df_together.columns[0]) #sort
            elif unit == "U2" and (name not in df2.iloc[:,0].values):
                colNames = df2.columns #get columns
                df_new = pd.DataFrame(data=newrow, columns=colNames) #Create new with columns
                df_together = pd.concat([df2,df_new],axis=0) #put together
                df2 = df_together.sort_values(df_together.columns[0]) #sort
            elif unit == "U3" and (name not in df3.iloc[:,0].values):
                colNames = df3.columns #get columns
                df_new = pd.DataFrame(data=newrow, columns=colNames) #Create new with columns
                df_together = pd.concat([df3,df_new],axis=0) #put together
                df3 = df_together.sort_values(df_together.columns[0]) #sort
            act[name] = [0,0,0,0,0,0,0,0,0,0,0,0,0,0]

    #Go to events
    def eventselect(self, eventselection: str):
        global event
        self.eventselect_label.set("Event")
        event = eventselection
        self.unitselection = Unitpicker()
        self.withdraw() #Hide Event Window
        self.unitselection.wm_protocol('WM_DELETE_WINDOW', lambda: self.eventwindow())

    #Move back a window
    def eventwindow(self):
        global event
        self.deiconify() #Bring event window up
        self.unitselection.deletion() #Destroy events window
        event = "" #Reset event name

    #Open new window
    def submitday(self, day: int):
        self.day_menu.set("Day")
        global df1, df2, df3, act
        together = {}

        for key in act:
            together[key] = [x * day for x in act[key]] #Times by day
        
        fin = {k: v for k, v in together.items() if v != ['','','','','','','','','','','','','','']}#keep only modified lines
        col_map = ["visits","music","outdoor","religion","food/drink","cards/games","arts/crafts","physical games","group activity",
                  "bus trips","special event","act on the go","noted","misc"]
        
        for k, v in fin.items(): #Get each new list
            for i, val in enumerate(v): #Go through each event
                if val: #If not empty
                    col_name = col_map[i] #get col name

                    if not df1.loc[df1['name'] == k].empty: #Find correct file
                        #If null add val, else add exisiting with val
                        if pd.isna(df1.loc[df1['name'] == k, col_name].values[0]):
                            df1.loc[df1['name'] == k, col_name] = f"{val}" 
                        else:
                            df1.loc[df1['name'] == k, col_name] = f"{df1.loc[df1['name'] == k,col_name].values[0]} {val}"
                            df1.loc[df1['name'] == k, col_name] = self.strings(df1.loc[df1['name'] == k,col_name].values[0])

                    if not df2.loc[df2['name'] == k].empty:
                        if pd.isna(df2.loc[df2['name'] == k, col_name].values[0]):
                            df2.loc[df2['name'] == k, col_name] = f"{val}"
                        else:
                            df2.loc[df2['name'] == k, col_name] = f"{df2.loc[df2['name'] == k,col_name].values[0]} {val}"
                            df2.loc[df2['name'] == k, col_name] = self.strings(df2.loc[df2['name'] == k,col_name].values[0])

                    if not df3.loc[df3['name'] == k].empty:
                        if pd.isna(df3.loc[df3['name'] == k, col_name].values[0]):
                            df3.loc[df3['name'] == k, col_name] = f"{val}"
                        else:
                            df3.loc[df3['name'] == k, col_name] = f"{df3.loc[df3['name'] == k,col_name].values[0]} {val}"
                            df3.loc[df3['name'] == k, col_name] = self.strings(df3.loc[df3['name'] == k,col_name].values[0])

        #Save files
        df1.to_csv('app_files/U1.csv',index=False)
        df2.to_csv('app_files/U2.csv',index=False)
        df3.to_csv('app_files/U3.csv',index=False)
        act = {} #Reset activity tracking

    #Organize dates
    def strings(self, s):
        num = s.split()
        uni = {int(float(x)) for x in num}
        sort = sorted(uni)
        return " ".join(map(str, sort))

    #Move data to excel file
    def submitmonth(self, month: str):
        global df1, df2, df3, act
        self.month_menu.set("Month")
        act = {}

        #Create file with each unit getting a seperate sheet
        with pd.ExcelWriter(f'{month}.xlsx', engine='xlsxwriter') as writer:
            df1.to_excel(writer, sheet_name='Unit1', index=False)
            df2.to_excel(writer, sheet_name='Unit2', index=False)
            df3.to_excel(writer, sheet_name='Unit3', index=False)

        #NEED TO DO THIS FOR EACH SHEET
        #Adding calculations to excel file
        wb = load_workbook(f"{month}.xlsx")
        ws = wb.active
        ws.insert_cols(1)
        ws.insert_rows(1)
        ws['A1']="=\"Total:\"&SUM(C1:P1)"
        ws['B1']="Total Per Activity"
        ws['A2']="Total Per Resident"
        #3-100
        #for row in range(3,76):
        #    ws[f"A{row}"] = f'=SUM(IF((C{row}:P{row})="",0,LEN(C{row}:P{row})-LEN(SUBSTITUTE(C{row}:P{row}," ",""))+1))'
        ws["A3"] = '=SUM(IF((C3:P3)="",0,LEN(C3:P3)-LEN(SUBSTITUTE(C3:P3," ",""))+1))'
        #a-z
        #for col in range(3, 17):
        #    col_letter = chr(64 + col)
        #    ws[f"{col_letter}1"] = f'=SUM(IF(({col_letter}3:{col_letter}75)="",0,LEN({col_letter}3:{col_letter}75)-LEN(SUBSTITUTE({col_letter}3:{col_letter}75," ",""))+1))'
        ws[f"C1"] = f'=SUM(IF((C3:C75)="",0,LEN(C3:C75)-LEN(SUBSTITUTE(C3:C75," ",""))+1))'
        wb.save(f"{month}.xlsx")

        #Reset csv files
        df1[["visits","music","outdoor","religion","food/drink","cards/games","arts/crafts","physical games","group activity",
                  "bus trips","special event","act on the go","noted","misc"]] = None
        df1.to_csv('app_files/U1.csv',index=False)
        df2[["visits","music","outdoor","religion","food/drink","cards/games","arts/crafts","physical games","group activity",
                  "bus trips","special event","act on the go","noted","misc"]] = None
        df2.to_csv('app_files/U2.csv',index=False)
        df3[["visits","music","outdoor","religion","food/drink","cards/games","arts/crafts","physical games","group activity",
                  "bus trips","special event","act on the go","noted","misc"]] = None
        df3.to_csv('app_files/U3.csv',index=False)

#Start the file
if __name__ == "__main__":
    app = App()
    app.mainloop()